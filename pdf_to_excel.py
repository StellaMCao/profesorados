import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

# Leer el PDF
pdf_path = r'C:\Users\stell\Downloads\correlativas (1).pdf'
pdf_file = open(pdf_path, 'rb')
reader = PyPDF2.PdfReader(pdf_file)

# Extraer todo el texto
full_text = ''
for page in reader.pages:
    full_text += page.extract_text() + '\n'

pdf_file.close()

# Crear el workbook de Excel
wb = Workbook()
ws = wb.active
ws.title = "Correlativas"

# Configurar estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezados
headers = ["Materia", "Correlativas / Requisitos", "Observaciones"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 30

# Procesar el texto y agregar datos
# Como el PDF parece tener información fragmentada, voy a agregar el texto completo
# para que puedas editarlo manualmente o podemos estructurarlo mejor si me das más detalles

row = 2
# Dividir por líneas y procesar
lines = [line.strip() for line in full_text.split('\n') if line.strip()]

# Agregar todo el contenido extraído para revisión
ws.cell(row=row, column=1, value="CONTENIDO EXTRAÍDO DEL PDF")
ws.cell(row=row, column=1).font = Font(bold=True)
row += 1

for line in lines:
    if line:
        ws.cell(row=row, column=1, value=line)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

# Guardar el archivo
output_path = r'C:\Users\stell\OneDrive\Escritorio\IA en educación\correlativas.xlsx'
wb.save(output_path)

print(f"Archivo Excel creado exitosamente en: {output_path}")
print(f"\nContenido extraído del PDF:")
print(full_text)
