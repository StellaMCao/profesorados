import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Leer el PDF
pdf_path = r'C:\Users\stell\Downloads\correlativas (1).pdf'
pdf_file = open(pdf_path, 'rb')
reader = PyPDF2.PdfReader(pdf_file)

# Extraer todo el texto por páginas
pages_text = []
for i, page in enumerate(reader.pages):
    text = page.extract_text()
    pages_text.append(f"=== PÁGINA {i+1} ===\n{text}")

full_text = '\n\n'.join(pages_text)
pdf_file.close()

# Crear el workbook de Excel con dos hojas
wb = Workbook()

# === HOJA 1: Contenido extraído del PDF ===
ws_raw = wb.active
ws_raw.title = "Contenido PDF"

# Estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezado
ws_raw.cell(row=1, column=1, value="CONTENIDO EXTRAÍDO DEL PDF - CORRELATIVAS")
ws_raw.cell(row=1, column=1).fill = header_fill
ws_raw.cell(row=1, column=1).font = header_font
ws_raw.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

# Agregar el texto completo
ws_raw.cell(row=2, column=1, value=full_text)
ws_raw.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical='top')
ws_raw.column_dimensions['A'].width = 120

# === HOJA 2: Plantilla estructurada para correlativas ===
ws_struct = wb.create_sheet("Correlativas Estructuradas")

# Encabezados
headers = ["Materia", "Año/Cuatrimestre", "Correlativas para Cursar", "Correlativas para Rendir", "Observaciones"]
for col, header in enumerate(headers, 1):
    cell = ws_struct.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Ajustar anchos de columna
ws_struct.column_dimensions['A'].width = 45  # Materia
ws_struct.column_dimensions['B'].width = 18  # Año/Cuatrimestre
ws_struct.column_dimensions['C'].width = 40  # Correlativas para Cursar
ws_struct.column_dimensions['D'].width = 40  # Correlativas para Rendir
ws_struct.column_dimensions['E'].width = 35  # Observaciones

# Agregar algunas filas de ejemplo para que puedas completar
ejemplo_materias = [
    "Ejemplo: Psicología General",
    "Ejemplo: Filosofía de la Psicología",
    "Ejemplo: Teoría y Epistemología Psicoanalítica",
    "",
    "",
    "",
]

for i, materia in enumerate(ejemplo_materias, 2):
    ws_struct.cell(row=i, column=1, value=materia)
    for col in range(1, 6):
        ws_struct.cell(row=i, column=col).border = border
        ws_struct.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical='top')

# Agregar filas vacías adicionales
for row in range(len(ejemplo_materias) + 2, 50):
    for col in range(1, 6):
        ws_struct.cell(row=row, column=col).border = border
        ws_struct.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

# Guardar el archivo
output_path = r'C:\Users\stell\OneDrive\Escritorio\IA en educación\correlativas.xlsx'
wb.save(output_path)

print(f"✓ Archivo Excel creado exitosamente en: {output_path}")
print(f"\nEl archivo contiene 2 hojas:")
print(f"  1. 'Contenido PDF' - Todo el texto extraído del PDF")
print(f"  2. 'Correlativas Estructuradas' - Plantilla para organizar las correlativas")
print(f"\nPuedes revisar el contenido del PDF en la primera hoja y luego")
print(f"estructurar la información en la segunda hoja.")
