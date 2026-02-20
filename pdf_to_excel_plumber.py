import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Definir rutas
pdf_path = r'C:\Users\stell\Downloads\correlativas (1).pdf'
output_path = r'C:\Users\stell\OneDrive\Escritorio\IA en educación\correlativas.xlsx'

# Crear workbook
wb = Workbook()

# === HOJA 1: Tablas Estructuradas ===
ws_tables = wb.active
ws_tables.title = "Correlativas"

# Estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
centered = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Procesar PDF
all_rows = []
full_text = ""

try:
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            # Extraer texto para la hoja de respaldo
            text = page.extract_text()
            if text:
                full_text += f"=== PÁGINA {i+1} ===\n{text}\n\n"
            
            # Extraer tablas
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    # Limpiar datos: quitar None y saltos de línea extra
                    clean_row = [cell.strip() if cell else "" for cell in row]
                    # Solo agregar si la fila tiene contenido relevante
                    if any(clean_row):
                        all_rows.append(clean_row)

    # Determinar el número máximo de columnas
    max_cols = max(len(row) for row in all_rows) if all_rows else 0
    
    if max_cols > 0:
        # Escribir encabezados genéricos si no los detectamos bien, o usar la primera fila
        # Asumiremos que la primera fila válida encontrada son los encabezados o usamos genéricos
        
        # Escribir datos
        current_row = 1
        for row_data in all_rows:
            # Escribir celdas
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws_tables.cell(row=current_row, column=col_idx, value=cell_value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = border
                
                # Si es la primera fila, darle estilo de encabezado
                if current_row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = centered
            
            current_row += 1
            
        # Ajustar anchos
        for col in range(1, max_cols + 1):
            ws_tables.column_dimensions[chr(64 + col)].width = 30
            
    else:
        ws_tables.cell(row=1, column=1, value="No se detectaron tablas claras. Ver hoja de Texto Completo.")

    # === HOJA 2: Texto Completo (Respaldo) ===
    ws_text = wb.create_sheet("Texto Completo")
    ws_text.cell(row=1, column=1, value=full_text)
    ws_text.column_dimensions['A'].width = 100
    ws_text.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_path)
    print(f"Éxito: Archivo creado en {output_path}")
    print(f"Filas extraídas: {len(all_rows)}")

except Exception as e:
    print(f"Error: {str(e)}")
